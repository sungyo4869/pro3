import openpyxl
import sys

# ファイルの中身を定義する関数
def create_table(n):
    wb = openpyxl.Workbook()
    sheet = workbook.active

    # ヘッダー行の作成
    for i in range(1, n+1):
        sheet.cell(row=1, column=i+1, value=i)

    # row = 1のとこ以外の計算と書き込み
    for i in range(1, n+1):
        sheet.cell(row=i+1, column=1, value=i)
        for j in range(1, n+1):
            sheet.cell(row=i+1, column=j+1, value=i*j)

    return wb

# ファイルを保存する関数
def save_file(wb, filename):
    wb.save(filename)
    print(f"Excelファイル '{filename}' が作成されました。")
    wb.close()

if __name__ == "__main__":
    # コマンドライン引数を指定しているかの判定
    if len(sys.argv) != 2:
        print("コマンドライン引数として数字 N を指定してください。")
        sys.exit(1)

    # コマンドライン引数をintにキャスト
    try:
        N = int(sys.argv[1])
    except ValueError:
        print("有効な数字を指定してください。")
        sys.exit(1)

    if N <= 0:
        print("正の整数を指定してください。")
        sys.exit(1)

    # ファイル名を定義
    filename = "multiplicationTable.xlsx"
    # 関数の呼び出し
    table = create_table(N)
    save_file(table, filename)
