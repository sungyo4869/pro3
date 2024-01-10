import openpyxl
import sys
from openpyxl.styles import Font

# ファイルの中身を定義する関数
def define_file_content(n):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet1 = wb.create_sheet(title="Sheet1", index=0)
    sheet2 = wb.create_sheet(title="Sheet2", index=1)
    sheet3 = wb.create_sheet(title="Sheet3", index=2)
    
    # ヘッダー行の作成
    
    create_table(sheet1, n)
    create_table(sheet2, n+1)
    create_table(sheet3, n+2)

    return wb

def create_table (sheet, n):
    for i in range(1, n+1):
        sheet.cell(row=1, column=i+1).font = Font(name = 'ＭＳ Ｐゴシック', bold=True)
        sheet.cell(row=1, column=i+1, value=i)
    # row = 1のとこ以外の計算と書き込み
    for i in range(1, n+1):
        sheet.cell(row=i+1, column=1).font = Font(name = 'ＭＳ Ｐゴシック', bold=True)
        sheet.cell(row=i+1, column=1, value=i)
        for j in range(1, n+1):
            sheet.cell(row=i+1, column=j+1, value=i*j)

# ファイルを保存する関数
def save_file(wb, filename):
    wb.save(filename)
    print(f"Excelファイル '{filename}' が作成されました。")
    wb.close()

if __name__ == "__main__":
    # コマンドライン引数を指定しているかの判定
    if len(sys.argv) != 2:
        print("コマンドライン引数として数字 N を指定してください。")
        sys.exit(1)

    # コマンドライン引数をintにキャスト
    try:
        N = int(sys.argv[1])
    except ValueError:
        print("有効な数字を指定してください。")
        sys.exit(1)

    if N <= 0:
        print("正の整数を指定してください。")
        sys.exit(1)

    # ファイル名を定義
    filename = "multiplicationTable.xlsx"
    # 関数の呼び出し
    sheets = define_file_content(N)
    
    save_file(sheets, filename)
