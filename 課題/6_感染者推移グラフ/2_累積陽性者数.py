import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib
from datetime import datetime, timedelta

excel_file_path = "C:\\Users\\mi211324\\Documents\\Project\\Python\\pro3\\演習課題\\課題\\6_感染者推移グラフ\\陽性者数.xlsx"
wb = openpyxl.load_workbook(excel_file_path)

sheet = wb["陽性者数(2021)"]

times = list()
people = list()
cumulative_value = list()

# 各セルの値を取得
for i in range(0, 222):
    # 日付を取得し、月/日のフォーマットに変換
    python_data = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + sheet.cell(row=i+1, column=1).value - 2)
    times.append(python_data.strftime("%m/%d"))
    
    people_per_day = sheet.cell(row=i+1, column=2).value
    
    # 感染者数を取得
    if i == 0:
        cumulative_value.append(people_per_day)
        continue
    
    # 先日の累積陽性者数と足し算する
    cumulative_value.append(cumulative_value[i - 1] + people_per_day)
    
    
plt.plot(times, cumulative_value)

# x軸に表示される値を調整、各月の1日を指定
selected_indices = [0,30,61,91,122,153,183,214]
selected_times = [times[i] for i in selected_indices]
plt.xticks(selected_times) 

plt.title("累積陽性者数(2021/4/1 - 2021/11/8)")

plt.xlabel("月日")
plt.ylabel("人数(人)")

plt.savefig("2.svg")