import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib
from datetime import datetime, timedelta

excel_file_path = "C:\\Users\\mi211324\\Documents\\Project\\Python\\pro3\\演習課題\\課題\\6_感染者推移グラフ\\陽性者数.xlsx"
wb = openpyxl.load_workbook(excel_file_path)

sheet = wb["陽性者数(2021)"]

times = list()
list_people = list()
cumulative_value = list()

# 各セルの値を取得
for i in range(184,215):
    # シートから値を取得
    people = sheet.cell(row=i, column=2).value
    time = sheet.cell(row=i, column=1).value
    
    # timeを月/日のフォーマットに変換し、timesへ格納
    python_data = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + time - 2)
    times.append(python_data.strftime("%m/%d"))
    
    # 感染者数を格納
    list_people.append(people)
    
    # 累積感染者数を格納
    if i == 184:
        cumulative_value.append(people)
        continue
    
    cumulative_value.append(cumulative_value[len(cumulative_value)-1] + people)
    
plt.plot(times, list_people)
plt.plot(times, cumulative_value)
# 凡例の追加
plt.legend(["感染者数", "累積陽性者数"])

# x軸に表示される値を調整、各月の1日を指定
selected_indices = [0,3,6,9,12,15,18,21,24,27,30]
selected_times = [times[i] for i in selected_indices]
plt.xticks(selected_times) 

plt.title("感染者数と累積陽性者数(2021年10月)")

plt.xlabel("月日")
plt.ylabel("人数(人)")

plt.savefig("3.svg")